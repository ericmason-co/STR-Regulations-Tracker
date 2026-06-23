"""Export jurisdictions.json to the master XLSX spreadsheet.

Rebuilds Global_STR_Regulations_Comprehensive_Database.xlsx with three sheets,
mirroring the original structure:
    1. Executive Summary  - counts by status / region, last refresh
    2. US - Detailed      - all US jurisdictions, 21 fields
    3. International       - all international jurisdictions, 21 fields
Plus a Changelog sheet driven by changelog.json.

Usage:
    python3 scripts/export_xlsx.py [output_path]
"""

from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from schema import FIELDS, load_changelog, load_jurisdictions

OUT_DEFAULT = (
    Path(__file__).resolve().parent.parent
    / "Global_STR_Regulations_Comprehensive_Database.xlsx"
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
STATUS_FILLS = {
    "Banned": "C0392B",
    "Restricted": "E67E22",
    "Active": "27AE60",
    "Pending": "F1C40F",
    "None": "95A5A6",
}


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _write_detail_sheet(ws, rows: list[dict]) -> None:
    headers = ["Region", "Country"] + [label for _, label in FIELDS]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for j in rows:
        line = [j.get("region", ""), j.get("country", "")]
        line += [str(j.get(key, "")) for key, _ in FIELDS]
        ws.append(line)
        status_cell = ws.cell(row=ws.max_row, column=5)  # 'Regulatory Status' col
        color = STATUS_FILLS.get(j.get("status", ""))
        if color:
            status_cell.fill = PatternFill("solid", fgColor=color)
            status_cell.font = Font(color="FFFFFF", bold=True)
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22


def _write_summary_sheet(ws, jurisdictions: list[dict], refresh: str) -> None:
    ws["A1"] = "Global STR Regulation Tracker - Executive Summary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Last full refresh: {refresh}"
    ws["A3"] = f"Total jurisdictions tracked: {len(jurisdictions)}"

    by_status = Counter(j.get("status", "Unknown") for j in jurisdictions)
    by_region = Counter(j.get("region", "Unknown") for j in jurisdictions)

    row = 5
    ws.cell(row=row, column=1, value="By Regulatory Status").font = Font(bold=True)
    row += 1
    for status, n in by_status.most_common():
        ws.cell(row=row, column=1, value=status)
        ws.cell(row=row, column=2, value=n)
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="By Region").font = Font(bold=True)
    row += 1
    for region, n in by_region.most_common():
        ws.cell(row=row, column=1, value=region)
        ws.cell(row=row, column=2, value=n)
        row += 1
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 12


def _write_changelog_sheet(ws, entries: list[dict]) -> None:
    headers = [
        "Date", "Jurisdiction", "Change Type", "Field",
        "Summary", "Effective Date", "Source", "Confidence",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for e in entries:
        ws.append([
            e.get("date", ""),
            e.get("jurisdiction_label", ""),
            e.get("change_type", ""),
            e.get("field", ""),
            e.get("summary", ""),
            e.get("effective_date", ""),
            e.get("source_url", ""),
            e.get("confidence", ""),
        ])
    widths = [12, 26, 14, 14, 60, 16, 40, 12]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def main(out_path: Path) -> None:
    data = load_jurisdictions()
    jurisdictions = data["jurisdictions"]
    refresh = data.get("meta", {}).get("last_full_refresh", "unknown")
    changelog = load_changelog()

    us = [j for j in jurisdictions if j.get("region") == "US"]
    intl = [j for j in jurisdictions if j.get("region") == "International"]

    wb = Workbook()
    _write_summary_sheet(wb.active, jurisdictions, refresh)
    wb.active.title = "Executive Summary"
    _write_detail_sheet(wb.create_sheet("US - Detailed"), us)
    _write_detail_sheet(wb.create_sheet("International"), intl)
    _write_changelog_sheet(wb.create_sheet("Changelog"), changelog["entries"])

    wb.save(out_path)
    print(f"Wrote {out_path} ({len(us)} US, {len(intl)} international).")


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else OUT_DEFAULT
    main(out)
